import nltk
import openpyxl
import pyphen as pyphen
from newspaper import Article, ArticleException
from nltk.corpus import opinion_lexicon
from nltk.corpus import stopwords
from nltk.tokenize import RegexpTokenizer
from nltk.tokenize import sent_tokenize

xlsx = openpyxl.load_workbook("/Users/simson/Downloads/input.xlsx")  # Insert the input file location
sheet = xlsx.active

output_sheet = openpyxl.Workbook()
op_sheet = output_sheet.active
op_sheet.title = "outputSheet"


def filterStopWords(text):
    stop_words = set(stopwords.words('english'))
    word_tokens = nltk.word_tokenize(text)
    filtered_sentence = []
    for w in word_tokens:
        if w not in stop_words:
            filtered_sentence.append(w)
    return filtered_sentence


def positive_score_of(filtered_text):
    positive_words = set(opinion_lexicon.positive())
    positive_score = 0
    for w in filtered_text:
        if w.lower() in positive_words:
            positive_score = positive_score + 1
    return positive_score


def negative_score_of(filtered_text):
    negative_words = set(opinion_lexicon.negative())
    negative_score = 0
    for w in filtered_text:
        if w.lower() in negative_words:
            negative_score = negative_score + 1
    return negative_score


def polarity_of(positive_score, negative_score):
    return (positive_score - negative_score) / (positive_score + negative_score) + 0.000001


def subjectivity_of(positive_score, negative_score, length):
    return (positive_score + negative_score) / (length + 0.000001)


def average_sentence_length_of(text):
    global i
    sent_tokenize_list = sent_tokenize(text)
    word_count = 0
    for i in range(0, len(sent_tokenize_list)):
        word_count += sent_tokenize_list[i].count(" ")
    return word_count / (i + 1)


def average_no_of_words_per_sentence_length_of(text):
    word_tokenize = nltk.word_tokenize(text)
    sent_tokenize_list = sent_tokenize(text)
    return len(word_tokenize) / len(sent_tokenize_list)


def word_count_of(text):
    tokenizer = RegexpTokenizer(r'\w+')
    return len(tokenizer.tokenize(text))


def complex_word_count_of(text, word_count):
    tokenizer = RegexpTokenizer(r'\w+')
    complex_word_count = 0
    total_syllable = 0
    word_tokenizer = tokenizer.tokenize(text)
    for i in range(0, len(word_tokenizer)):
        dic = pyphen.Pyphen(lang='en')
        count = dic.inserted(word_tokenizer[i]).count("-") + 1
        total_syllable = total_syllable + dic.inserted(word_tokenizer[i]).count("-") + 1
        if count > 2:
            complex_word_count = complex_word_count + 1
    syllable_per_word = total_syllable / word_count
    return complex_word_count, syllable_per_word


def personal_pronoun_of(text):
    word_tokens = nltk.word_tokenize(text)
    tagged = nltk.pos_tag(word_tokens)
    prp_tag = [t for t in tagged if t[1] == "PRP"]
    return len(prp_tag)


def average_word_length_of(text, word_count):
    tokenizer = RegexpTokenizer(r'\w+')
    word_tokens = tokenizer.tokenize(text)
    character_length = 0
    for i in range(0, len(word_tokens)):
        character_length = character_length + len(word_tokens[i])

    return character_length / word_count


def sentimentAnalysis(text, output_cell_address):
    filtered_text = filterStopWords(text)
    positive_score = positive_score_of(filtered_text)
    negative_score = negative_score_of(filtered_text)
    polarity_score = polarity_of(positive_score, negative_score)
    subjectivity_score = subjectivity_of(positive_score, negative_score, len(filtered_text))
    average_sentence_length = average_sentence_length_of(text)
    word_count = word_count_of(text)
    complex_word_count, syllable_per_word = complex_word_count_of(text, word_count)
    percentage_of_complex_words = complex_word_count / word_count
    average_no_of_words_per_sentence_length = average_no_of_words_per_sentence_length_of(text)
    fog_index = (average_sentence_length + percentage_of_complex_words) * 0.4
    personal_pronoun_count = personal_pronoun_of(text)
    average_word_length = average_word_length_of(text, word_count)
    print(f"Positive Score: {positive_score}")
    op_sheet.cell(output_cell_address, 3).value = positive_score
    print(f"Negative Score: {negative_score}")
    op_sheet.cell(output_cell_address, 4).value = negative_score
    print(f"Polarity Score :{polarity_score}")
    op_sheet.cell(output_cell_address, 5).value = polarity_score
    print(f"Subjectivity Score :{subjectivity_score}")
    op_sheet.cell(output_cell_address, 6).value = subjectivity_score
    print(f"Average sentence length :{average_sentence_length}")
    op_sheet.cell(output_cell_address, 7).value = average_sentence_length
    print(f"Percentage of Complex Word Count: {percentage_of_complex_words * 100} %")
    op_sheet.cell(output_cell_address, 8).value = percentage_of_complex_words
    print(f"Fog Index: {fog_index}")
    op_sheet.cell(output_cell_address, 9).value = fog_index
    print(f"Average words per sentence length :{average_no_of_words_per_sentence_length}")
    op_sheet.cell(output_cell_address, 10).value = average_no_of_words_per_sentence_length
    print(f"Complex Word Count: {complex_word_count}")
    op_sheet.cell(output_cell_address, 11).value = complex_word_count
    print(f"Word Count: {word_count}")
    op_sheet.cell(output_cell_address, 12).value = word_count
    print(f"Syllable per word : {syllable_per_word}")
    op_sheet.cell(output_cell_address, 13).value = syllable_per_word
    print(f"Personal Pronouns Count: {personal_pronoun_count}")
    op_sheet.cell(output_cell_address, 14).value = personal_pronoun_count
    print(f"Average Word Length : {average_word_length}")
    op_sheet.cell(output_cell_address, 15).value = average_word_length


def output_formatting():
    op_sheet.cell(1, 1).value = "URL_ID"
    op_sheet.cell(1, 2).value = "URL"
    op_sheet.cell(1, 3).value = "POSITIVE SCORE"
    op_sheet.cell(1, 4).value = "NEGATIVE SCORE"
    op_sheet.cell(1, 5).value = "POLARITY SCORE"
    op_sheet.cell(1, 6).value = "SUBJECTIVITY SCORE"
    op_sheet.cell(1, 7).value = "AVG SENTENCE LENGTH"
    op_sheet.cell(1, 8).value = "PERCENTAGE OF COMPLEX WORDS"
    op_sheet.cell(1, 9).value = "FOG INDEX"
    op_sheet.cell(1, 10).value = "AVG NUMBER OF WORDS PER SENTENCE"
    op_sheet.cell(1, 11).value = "COMPLEX WORD COUNT"
    op_sheet.cell(1, 12).value = "WORD COUNT"
    op_sheet.cell(1, 13).value = "SYLLABLE PER WORD"
    op_sheet.cell(1, 14).value = "PERSONAL PRONOUNS"
    op_sheet.cell(1, 15).value = "AVG WORD LENGTH"


def articleCleanup(urlSample, cell_address):
    try:
        print(urlSample)
        article = Article(urlSample)
        article.download()  # Downloads the article from the web
        article.parse()  # Cleans up the article by removing HTML tags
        article.nlp()  # Prepares the article for natural language processing
        text = article.text
        sentimentAnalysis(text, cell_address)
    except ArticleException as ae:
        print(ae)
        op_sheet.cell(cell_address, 1).value = sheet.cell(cell_address, 1).value
        op_sheet.cell(cell_address, 2).value = urlSample
        for op_i in range(3, 16):
            op_sheet.cell(i, op_i).value = "404 Error"


for i in range(2, sheet.max_row + 1):
    output_formatting()
    print(i)
    url = sheet.cell(i, 2).value
    op_sheet.cell(i, 1).value = sheet.cell(i, 1).value
    op_sheet.cell(i, 2).value = url
    articleCleanup(url, i)

output_sheet.save("/Users/simson/Downloads/output_sheet_assignment.xlsx")
